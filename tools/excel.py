import datetime

import openpyxl

LESSONS_DAILY = {
    1: '8:00-9:30',
    2: '9:40-11:10',
    3: '11:50-13:20',
    4: '13:40-15:10',
    5: '15:20-16:50',
    6: '17:00-18:30',
    7: '18:40-20:10',
}

LESSONS_SATURDAY = {
    1: '8:00-9:30',
    2: '9:40-11:10',
    3: '11:30-13:00',
    4: '13:10-14:40',
    5: '14:50-16:20',
    6: '16:30-18:00',
    7: '18:00-19:30'
}


def parse_groups(cells, pair, weekday):
    for cell_obj in cells:
        if weekday == 5:
            string = '•' + LESSONS_SATURDAY[pair] + ' - '
        else:
            string = '•' + LESSONS_DAILY[pair] + ' - '
        string += ' - '.join([str(cell.value).replace('\n', '|').replace(' /', '/') for cell in cell_obj])
        string = string.replace(' - None', '')
    return string


def name_to_standard(name) -> str:
    if not name or all([i == '_' for i in name]):
        return ''
    name = name.split()
    if len(name) == 2:
        surname = name[0]
        first_name = '.'.join([i for i in name[1] if i.isupper()]) + '.'
        return surname + ' ' + first_name
    elif len(name) == 1:
        index = None
        surname, c = name[0], 1
        for i in surname[1:]:
            if i.isupper():
                index = c
                break
            c += 1
        if index:
            first_name = '.'.join([i for i in surname[index:] if i.isupper()]) + '.'
            return surname[:index] + ' ' + first_name
        return ''


def parse_teachers(cells, pair, weekday, group):
    t_pairs = {}
    teachers = cells[0][2].value
    if not teachers:
        return {}
    c = 0
    for teacher_pairs in teachers.split('\n'):
        pairs = cells[0][1].value.split('\n')
        for teacher in (teacher_pairs.split('/') if '/' in teacher_pairs else teacher_pairs.split('\\')):
            teacher = name_to_standard(teacher.strip())
            t_pairs[teacher.strip()] = t_pairs.get(teacher, {})
            t_pairs[teacher.strip()][weekday] = []
            if len(pairs) == len(teachers.split('\n')) > 1:
                t_pairs[teacher.strip()][weekday].append(
                    '•' + (LESSONS_DAILY[pair] if weekday != 5 else LESSONS_SATURDAY[pair]) + ' - Группа ' + str(
                        group) + ' - Аудитория ' + '|'.join(str(cells[0][0].value).split('\n')) + ' - ' + pairs[
                        c].strip() + (' - До пересменки' if c == 0 else ' - После пересменки'))
            else:
                t_pairs[teacher.strip()][weekday].append(
                    '•' + (LESSONS_DAILY[pair] if weekday != 5 else LESSONS_SATURDAY[pair]) + ' - Группа ' + str(
                        group) + ' - Аудитория ' + '/'.join(str(cells[0][0].value).split('\n')) + ' - ' + pairs[
                        0].strip())
        c += 1
    return t_pairs


def parse_workbook(workbook, workbook_number):
    sheet_names = workbook.sheetnames
    groups = {}
    teachers = {}
    for sheet_number in range(len(sheet_names)):  # Цикл по листам
        start_col, end_col = ord('C'), ord('E')
        sheet = workbook[sheet_names[sheet_number]]
        for group_number in sheet_names[sheet_number].split(','):  # Цикл про группам
            groups[group_number] = {}
            if workbook_number in (1, 2):
                start_row = 8 if sheet_number == 0 else 7
            else:
                start_row = 7
            for weekday in range(6):  # 6 дней недели
                pair = 1
                groups[group_number][weekday] = ''
                for rows in range(start_row, start_row + (7 if workbook_number == 4 else 6)):  # Берем по 6 пар
                    cells = sheet[f'{chr(start_col)}{rows}:{chr(end_col)}{rows}']
                    groups[group_number][weekday] += parse_groups(cells, pair, weekday) + '\n\n'
                    t_pairs = parse_teachers(cells, pair, weekday, group_number)
                    for name in t_pairs:
                        teachers[name] = teachers.get(name, {})
                        teachers[name][weekday] = teachers[name].get(weekday, [])
                        for p in t_pairs[name][weekday]:
                            teachers[name][weekday].append(p + '\n')
                    pair += 1
                start_row += 6
            start_col = end_col + 1
            end_col = start_col + 2
    return groups, teachers


def get_temp_timetable_dates(workbook):
    dates = []
    cells = workbook[workbook.sheetnames[0]][f'A{8}:A{38}']
    for cellObj in cells:
        for cell in cellObj:
            if cell.value:
                dates.append(cell.value.split('\n')[-1].split('.'))
    return dates


def main(temp: bool = False):
    groups = {}
    teachers = {}
    temp_dates = None
    for i in range(1, 5):
        workbook = openpyxl.load_workbook(
            f'workbooks/rasp{i}.xlsx' if not temp else f'workbooks/temp_rasp{i}.xlsx')
        if temp and i == 1:
            dates = get_temp_timetable_dates(workbook)
            dates = list(map(lambda x: datetime.date(*map(int, x[::-1])), dates))
            if (max(dates) - datetime.date.today()).days < 0:
                print('Временное расписание уже устарело')
                return False
            temp_dates = dates
        group_dict, teacher_dict = parse_workbook(workbook, i)
        groups.update(group_dict)
        for teacher in teacher_dict:
            for weekday in teacher_dict[teacher]:
                teachers[teacher] = teachers.get(teacher, {})
                teachers[teacher][weekday] = teachers[teacher].get(weekday, []) + teacher_dict[teacher][weekday]
    if not temp_dates:
        return groups, teachers
    return groups, teachers, temp_dates
